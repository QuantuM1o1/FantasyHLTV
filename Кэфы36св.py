import time
from openpyxl import load_workbook
wb = load_workbook("./Кэфы.xlsx", data_only=True)
sheet = wb.get_sheet_by_name('2022 (36)')
l12 = int(input())
r3 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r3[it] = sheet.cell(row = i, column = 4).value
    it += 1
r6 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r6[it] = sheet.cell(row = i, column = 5).value
    it += 1
f = [0] * l12
it = 0
for i in range(2, l12 + 2):
    f[it] = sheet.cell(row = i, column = 6).value
    it += 1
sras = 0
gsras = 1000
shag = 2
drob = 10 * shag
d = 15 * shag
for i in range(d):
    inorm = i / drob
    for j in range(d):
        jnorm = j / drob
        for k in range(2 * d):
            knorm = (k - d) / drob
            for l in range(l12):
                p = r3[l] * inorm + r6[l] * jnorm + knorm
                ras = pow(p - f[l], 2)
                sras += ras
            if gsras > sras:
                gsras = sras
                gi = inorm
                gj = jnorm
                gk = knorm
            sras = 0
f2 = open('kfin.txt', 'w')
f2.write(str(round(gsras, 2)))
f2.write('\n')
f2.write('3 месяц =' + str(gi))
f2.write('\n')
f2.write('6 месяцев =' + str(gj))
f2.write('\n')
f2.write('Свободный =' + str(gk))
f2.write('\n')
f2.close()        
