import time
from openpyxl import load_workbook
wb = load_workbook("./Кэфы.xlsx", data_only=True)
sheet = wb.get_sheet_by_name('2022 (136)')
l12 = int(input())
r1 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r1[it] = sheet.cell(row = i, column = 3).value
    it += 1
r3 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r3[it] = sheet.cell(row = i, column = 4).value
    it += 1
r6 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r6[it] = sheet.cell(row = i, column = 5).value
    it += 1
f = [0] * l12
it = 0
for i in range(2, l12 + 2):
    f[it] = sheet.cell(row = i, column = 6).value
    it += 1
sras = 0
gsras = 1000
shag = 2
drob = 10 * shag
d = 15 * shag
t0 = time.process_time()
for i in range(d):
    inorm = i / drob
    for j in range(d):
        jnorm = j / drob
        for k in range(d):
            knorm = k / drob
            for l in range(2 * d):
                lnorm = (l - d) / drob
                for m in range(l12):
                    p = r1[m] * inorm + r3[m] * jnorm + r6[m] * knorm + lnorm
                    ras = pow(p - f[m], 2)
                    sras += ras
                if gsras > sras:
                    gsras = sras
                    gi = inorm
                    gj = jnorm
                    gk = knorm
                    gl = lnorm
                sras = 0
t1 = time.process_time() - t0
print(t1-t0)
f2 = open('kfin.txt', 'w')
f2.write(str(round(gsras, 2)))
f2.write('\n')
f2.write('1 месяц =' + str(gi))
f2.write('\n')
f2.write('3 месяц =' + str(gj))
f2.write('\n')
f2.write('6 месяцев =' + str(gk))
f2.write('\n')
f2.write('Свободный =' + str(gl))
f2.write('\n')
f2.close()        
