f1 = open('st.txt','r')
exname = f1.read()
f1.close()
from openpyxl import load_workbook
wb = load_workbook(exname, data_only=True)
sheet = wb.get_sheet_by_name('Статистика')
lenghtfind = [0]
i1 = 2
while sheet.cell(row = i1, column = 2).value != None:
     lenghtfind.append(sheet.cell(row = i1, column = 2).value)
     i1 += 1
l12 = len(lenghtfind) - 1
pname = [0] * l12
it = 0
for i in range(2, l12 + 2):
    pname[it] = sheet.cell(row = i, column = 2).value
    it += 1
tname = [0] * l12
it = 0
for i in range(2, l12 + 2):
    tname[it] = sheet.cell(row = i, column = 1).value
    it += 1
rate = [0] * l12
it = 0
for i in range(2, l12 + 2):
    rate[it] = (sheet.cell(row = i, column = 3).value) ** 2
    it += 1
it = 0
price = [0] * l12
for i in range(2, l12 + 2):
    price[it] = sheet.cell(row = i, column = 4).value
    it += 1
gs = 0
p = 0
gp = 1001
gc = [0,0,0,0,0]
s = 0
for i in range(l12):
    s = rate[i]
    p = price[i]
    for j in range(i,l12):
        if j==i:
            continue
        s += rate[j]
        p += price[j]
        for k in range(j,l12):
            if k==j:
                continue
            if k==i:
                continue
            if tname[k] == tname[j] == tname[i]:
                continue
            s += rate[k]
            p += price[k]
            for l in range(k,l12):
                if l==k:
                    continue
                if l==j:
                    continue
                if l==i:
                    continue
                if tname[l] == tname [k] == tname[j] or tname[l] == tname[k] == tname[i] or tname[l] == tname [j] == tname[i]:
                    continue
                s += rate[l]
                p += price[l]
                for m in range(l,l12):
                    if m==l:
                        continue
                    if m==k:
                        continue
                    if m==j:
                        continue
                    if m==i:
                        continue
                    if tname[m] == tname[l] == tname[k] or tname[m] == tname[l] == tname[j] or tname[m] == tname[l] == tname[i] or tname[m] == tname[k] == tname[j] or tname[m] == tname[k] == tname[i] or tname[m] == tname[j] == tname[i]:
                        continue
                    s += rate[m]
                    p += price[m]
                    if p < gp and s > gs:
                        gs = s
                        gc = [i,j,k,l,m]
                    s -= rate[m]
                    p -= price[m]
                s -= rate[l]
                p -= price[l]
            s -= rate[k]
            p -= price[k]
        s -= rate[j]
        p -= price[j]
    s -= rate[i]
    p -= price[i]
f2 = open('ifin.txt', 'w')
f2.write(str(round(gs, 2)))
f2.write('\n')
for ita in range(5):
    f2.write(pname[gc[ita]])
    f2.write('\n')
f2.close()    
