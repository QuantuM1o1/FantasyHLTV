f1 = open('st.txt','r')
exname = f1.read()
f1.close()
from openpyxl import load_workbook
wb = load_workbook(exname, data_only=True)
sheet = wb.get_sheet_by_name('Итоги')
l12 = int(input())
r1 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r1[it] = sheet.cell(row = i, column = 3).value
    it += 1
r3 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r3[it] = sheet.cell(row = i, column = 4).value
    it += 1
r6 = [0] * l12
it = 0
for i in range(2, l12 + 2):
    r6[it] = sheet.cell(row = i, column = 5).value
    it += 1
f = [0] * l12
it = 0
for i in range(2, l12 + 2):
    f[it] = sheet.cell(row = i, column = 6).value
    it += 1
sras = 0
gsras = 1000
for i in range(151):
    inorm = i / 100
    for j in range(i+1):
        jnorm = j / 100
        for k in range(l12):
            p = r1[k] * jnorm + r3[k] * (inorm - jnorm)
            ras = pow(p - f[k], 2)
            sras += ras
        if gsras > sras:
            gsras = sras
            gi = inorm
            gj = jnorm
        sras = 0
f2 = open('kfin.txt', 'w')
f2.write(str(round(gsras, 2)))
f2.write('\n')
f2.write('Суммарно =' +  str(gi))
f2.write('\n')
f2.write('1 месяц =' + str(gj))
f2.write('\n')
f2.close()        
